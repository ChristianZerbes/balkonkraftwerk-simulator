#!/usr/bin/env python3
"""
szenarien_pivot_aktualisieren.py
────────────────────────────────
Liest result.csv und erzeugt szenarien_pivot.xlsx im gleichen Ordner.

Ausführen:
  - Doppelklick (Windows: ggf. "Als Administrator ausführen" nicht nötig)
  - oder Terminal: python szenarien_pivot_aktualisieren.py

Voraussetzung (einmalig):
  pip install pandas openpyxl
"""

import sys, os

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Fehlende Pakete. Bitte einmalig ausführen:")
    print("  pip install pandas openpyxl")
    input("\nEnter drücken zum Beenden...")
    sys.exit(1)

file_path = os.path.join("../../result", "result.csv")


csv_path   = os.path.join("../../result", "result.csv")
xlsx_path  = os.path.join("../../result", "szenarien_pivot.xlsx")

if not os.path.exists(csv_path):
    print(f"FEHLER: result.csv nicht gefunden in:\n  {script_dir}")
    input("\nEnter drücken zum Beenden...")
    sys.exit(1)

print(f"Lese: {csv_path}")
df = pd.read_csv(csv_path, sep=";", decimal=",")

# Szenarien-Spalten: alle außer datum, normiert, delta und *_delta
scenario_cols = [c for c in df.columns
                 if c not in ("datum", "normiert", "delta")
                 and not c.endswith("_delta")]

rows = []
for _, r in df.iterrows():
    for sz in scenario_cols:
        # "real" verwendet die Spalte "delta"; alle anderen "szXX_delta"
        delta_key = "delta" if sz == "real" else f"{sz}_delta"
        rows.append({
            "Datum":     r["datum"],
            "Szenario":  sz,
            "Verbrauch": float(str(r[sz]).replace(",", ".")),
            "Delta":     float(str(r[delta_key]).replace(",", ".")),
        })

result = (pd.DataFrame(rows)
          .sort_values(["Datum", "Szenario"])
          .reset_index(drop=True))

print(f"{len(result)} Zeilen erzeugt  "
      f"({len(df)} Datumseinträge × {len(scenario_cols)} Szenarien: {', '.join(scenario_cols)})")

# ── Excel aufbauen ─────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Daten"

hdr_fill = PatternFill("solid", start_color="1F4E79")
hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
dat_font = Font(name="Arial", size=10)
alt_fill = PatternFill("solid", start_color="EBF3FB")
thin     = Side(style="thin", color="CCCCCC")
border   = Border(left=thin, right=thin, top=thin, bottom=thin)

headers    = ["Datum", "Szenario", "Verbrauch", "Delta"]
col_widths = [14, 12, 12, 12]

for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")
    c.border = border
    ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]

for row_idx, row in result.iterrows():
    er   = row_idx + 2
    vals = [row["Datum"], row["Szenario"], row["Verbrauch"], row["Delta"]]
    fill = alt_fill if row_idx % 2 == 0 else None
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=er, column=ci, value=val)
        c.font   = dat_font
        c.border = border
        if fill:
            c.fill = fill
        if ci in (3, 4):
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right")
        else:
            c.alignment = Alignment(horizontal="center")

ws.freeze_panes    = "A2"
ws.auto_filter.ref = f"A1:D{len(result) + 1}"

wi = wb.create_sheet("Info")
wi["A1"] = "Aktualisierung"
wi["A1"].font = Font(bold=True, name="Arial", size=12)
wi["A2"] = (
    "Dieses Skript (szenarien_pivot_aktualisieren.py) im gleichen Ordner "
    "wie result.csv ausführen, um szenarien_pivot.xlsx neu zu erzeugen.\n\n"
    "Neue Zeilen in result.csv werden automatisch berücksichtigt."
)
wi["A2"].font = Font(name="Arial", size=10)
wi["A2"].alignment = Alignment(wrap_text=True)
wi.column_dimensions["A"].width = 65
wi.row_dimensions[2].height = 60

wb.save(xlsx_path)
print(f"\nGespeichert: {xlsx_path}")
input("\nFertig! Enter drücken zum Schließen...")
